# Imports #
from selenium import webdriver
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.common.by import By
from selenium.webdriver.common.action_chains import ActionChains
from selenium.webdriver.remote.webelement import WebElement
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
import time
import os
from os import path
from openpyxl import load_workbook
from webdriver_manager.chrome import ChromeDriverManager


script_dir = os.path.dirname(os.path.abspath(__file__))

# Drop images logic
JS_DROP_FILES = """
var k=arguments,d=k[0],g=k[1],c=k[2],m=d.ownerDocument||document;for(var e=0;;){var f=d.getBoundingClientRect(),b=f.left+(g||(f.width/2)),a=f.top+(c||(f.height/2)),h=m.elementFromPoint(b,a);
if(h&&d.contains(h)){break}if(++e>1){var j=new Error('Element not interactable');j.code=15;throw j}d.scrollIntoView({behavior:'instant',block:'center',inline:'center'})}var l=m.createElement('INPUT');
l.setAttribute('type','file');l.setAttribute('multiple','');l.setAttribute('style','position:fixed;z-index:2147483647;left:0;top:0;');l.onchange=function(q){l.parentElement.removeChild(l);q.stopPropagation();
var r={constructor:DataTransfer,effectAllowed:'all',dropEffect:'none',types:['Files'],files:l.files,setData:function u(){},getData:function o(){},clearData:function s(){},setDragImage:function i(){}};
if(window.DataTransferItemList){r.items=Object.setPrototypeOf(Array.prototype.map.call(l.files,function(x){return{constructor:DataTransferItem,kind:'file',type:x.type,getAsFile:function v(){return x},getAsString:function y(A){var z=new FileReader();z.onload=function(B){A(B.target.result)};
z.readAsText(x)},webkitGetAsEntry:function w(){return{constructor:FileSystemFileEntry,name:x.name,fullPath:'/'+x.name,isFile:true,isDirectory:false,file:function z(A){A(x)}}}}}),{constructor:DataTransferItemList,add:function t(){},clear:function p(){},remove:function n(){}})}['dragenter','dragover','drop'].forEach(function(v){var w=m.createEvent('DragEvent');
w.initMouseEvent(v,true,true,m.defaultView,0,0,0,b,a,false,false,false,false,0,null);Object.setPrototypeOf(w,null);w.dataTransfer=r;Object.setPrototypeOf(w,DragEvent.prototype);h.dispatchEvent(w)})};m.documentElement.appendChild(l);l.getBoundingClientRect();return l
"""


def drop_files(element, files, offsetX=0, offsetY=0):
    driver = element.parent
    isLocal = not driver._is_remote or "127.0.0.1" in driver.command_executor._url
    paths = []
    for file in files if isinstance(files, list) else [files]:
        if not path.isfile(file):
            raise FileNotFoundError(file)
        paths.append(file if isLocal else element._upload(file))

    value = "\n".join(paths)
    elm_input = driver.execute_script(JS_DROP_FILES, element, offsetX, offsetY)
    elm_input._execute("sendKeysToElement", {"value": [value], "text": value})


# Important to link it to DOM
WebElement.drop_files = drop_files


# Read data from excel
def read_excel_tabs(file_path):
    workbook = load_workbook(file_path)
    sheet = workbook.active
    list_of_dicts = []

    for row in sheet.iter_rows(min_row=2, values_only=True):
        image, title, price, category, condition, color, description, location = row
        if location:
            entry_dict = {
                "Image": image,
                "Title": title,
                "Price": price,
                "Category": category,
                "Condition": condition,
                "Color": color,
                "Description": description,
                "Location": location,
            }
            list_of_dicts.append(entry_dict)
    return list_of_dicts


# Read data from excel
def read_profiles(file_path):
    workbook = load_workbook(file_path)
    sheet = workbook.active
    list_of_dicts = []

    for row in sheet.iter_rows(min_row=2, values_only=True):
        email, password = row
        if email:
            entry_dict = {"Email": email, "Password": password}
            list_of_dicts.append(entry_dict)
    return list_of_dicts


# Check if all images exist
def images_exist(data):
    i = 0
    no_image_missing = True
    for tab in data:
        i += 1
        if not path.exists(os.path.join(script_dir, 'images', tab["Image"])):
            no_image_missing = False
            print("Tab# " + str(i) + " : Image not found : " + tab["Image"])
    return no_image_missing


# Open Chrome using webdriver-manager
def open_browser():
    # Set up Chrome options
    options = Options()
    options.add_argument("--disable-notifications")  # Disable notifications
    options.add_argument("--no-sandbox")
    options.add_argument("--disable-dev-shm-usage")

    # Automatically download and set up the correct ChromeDriver
    service = Service(ChromeDriverManager().install())
    
    # Initialize the Chrome WebDriver with options
    driver = webdriver.Chrome(service=service, options=options)
    
    return driver

# Read settings from settings file
def read_settings():
    data = open("config.txt", "r").read().split("\n")
    config = {}
    for line in data:
        if line == "" or line.startswith("#"):
            continue
        line = line.replace("'", "")
        key, value = line.split(" = ")
        config[key] = value
    return config


# Log into facebook using creds
def login_facebook(driver, actions, profile):
    driver.find_element(By.ID, "email").click()
    actions.send_keys(profile["Email"]).perform()
    time.sleep(0.5)
    driver.find_element(By.ID, "pass").click()
    actions.send_keys(profile["Password"]).perform()
    time.sleep(1)
    driver.find_element(
        By.XPATH, "//button[contains(@data-testid,'royal_login_button')]"
    ).click()


# Open Tabs of new listing
def open_tabs(driver, count, tab):
    if tab == "1":  # item = 1 , tab = 1
        url = "https://www.facebook.com/marketplace/create/item/"
    else:  
        url = "https://www.facebook.com/marketplace/you/selling"

    # Open new tabs
    driver.get(url)
    for i in range(1, count):
        driver.execute_script("window.open('{}', '_blank');".format(url))


# Check if condition is same
def is_same_condition(text, tab_text):
    text = text.split()
    tab_text = tab_text.split()
    if text[0] == tab_text[0] and text[-1] == tab_text[-1]:
        return True
    return False


# Submit as fast as possible
def submit_quickly(driver):
    i = 0
    for handle in driver.window_handles:
        i += 1
        driver.switch_to.window(handle)
        # Publish
        try:
            publish_button = driver.find_element(
                By.XPATH, "//div[contains(@aria-label,'Publish')]"
            )
            driver.execute_script("arguments[0].scrollIntoView(true);", publish_button)
            time.sleep(0.2)
            publish_button.click()
        except:
            try:
                # Focus as well
                publish_button = driver.find_element(
                    By.XPATH, "//div[contains(@aria-label,'Publish')]"
                )
                driver.execute_script(
                    "arguments[0].scrollIntoView(true);", publish_button
                )
                driver.execute_script(
                    "arguments[0].setAttribute('aria-hidden', 'false')", publish_button
                )
                driver.execute_script(
                    "arguments[0].setAttribute('tabindex', '0')", publish_button
                )
                driver.execute_script("arguments[0].focus();", publish_button)
                driver.execute_script("arguments[0].click();", publish_button)
                actions.move_to_element(next_button).click(next_button).perform()
                time.sleep(0.2)
                publish_button.click()
            except:
                pass
        print("Published Tab#", i)


# Fill content
def fill_data(driver, actions, tab):
    # Obtiene el directorio del script
    script_dir = os.path.dirname(os.path.abspath(__file__))
    print(f"Directorio del script: {script_dir}")

    # Construye la ruta completa a la imagen en la carpeta 'images'
    image_path = os.path.join(script_dir, 'images', tab["Image"])
    print(f"Ruta de la imagen: {image_path}")

    # Encuentra los botones
    buttons = driver.find_elements(By.XPATH, "//div[contains(@role,'button')]")
    print(f"Botones encontrados: {len(buttons)}")

    image_drop = None
    more_details_element = None

    for button in buttons:
        try:
            # Busca el texto en los elementos hijos del div
            children_texts = [child.text.lower() for child in button.find_elements(By.XPATH, './/*')]
            print(f"Textos de los hijos del botón: {children_texts}")

            if any(
                "añadir fotos" in text
                and "o arrastra y suelta" in text
                for text in children_texts
            ):
                image_drop = button
                print("Botón de 'drag and drop' encontrado.")
            if any(
                "more details" in text
                for text in children_texts
            ):
                more_details_element = button
                print("Botón de 'more details' encontrado.")

        except Exception as e:
            print(f"Error al procesar el botón: {e}")

    if image_drop:
        # Usa el path de la imagen para hacer el drag and drop
        print("Realizando drag and drop de la imagen.")
        image_drop.drop_files(image_path)
    else:
        print("No se encontró el botón de 'drag and drop'")

    # Title
    print("Configurando el título.")
    title_element = driver.find_element(
        By.XPATH, "//label[contains(@aria-label,'Título')]"
    )
    driver.execute_script("arguments[0].scrollIntoView(true);", title_element)
    time.sleep(0.2)
    title_element.click()
    actions.send_keys(tab["Title"]).perform()

    # Price
    print("Configurando el precio.")
    driver.find_element(By.XPATH, "//label[contains(@aria-label,'Precio')]").click()
    actions.send_keys(tab["Price"]).perform()

    # Category
    print("Configurando la categoría.")
    cat_elem = driver.find_element(
        By.XPATH, "//label[contains(@aria-label,'Categoría')]"
    )
    driver.execute_script("arguments[0].scrollIntoView(true);", cat_elem)
    time.sleep(0.2)
    cat_elem.click()
    time.sleep(3)
    elements = driver.find_elements(By.XPATH, "//div[contains(@role,'button')]")
    for elem in elements:
        try:
            if str(elem.text.strip()).lower() == tab["Category"].strip().lower():
                elem.click()
                print(f"Categoría seleccionada: {tab['Category']}")
                break
        except:
            pass

    # Condition
    print("Configurando el estado.")
    con_elem = driver.find_element(
        By.XPATH, "//label[contains(@aria-label,'Estado')]"
    )
    con_elem.click()
    time.sleep(3)
    elements = driver.find_elements(By.XPATH, "//div[contains(@role,'option')]")
    for elem in elements:
        try:
            if is_same_condition(
                str(elem.text.strip()).lower(), tab["Condition"].strip().lower()
            ):
                elem.click()
                print(f"Estado seleccionado: {tab['Condition']}")
                break
        except:
            pass
    time.sleep(1)
    # Color
    if tab["Color"] != "" and tab["Color"] is not None:
        print("Configurando el color.")
        color_element = driver.find_element(
            By.XPATH, "//label[contains(@aria-label,'Color')]"
        )
        driver.execute_script("arguments[0].scrollIntoView(true);", color_element)
        color_element.click()
        actions.send_keys(tab["Color"]).perform()

    # Description
    if tab["Description"] != "" and tab["Description"] is not None:
        print("Configurando la descripción.")
        desc_element = driver.find_element(
            By.XPATH, "//label[contains(@aria-label,'Descripción')]"
        )
        driver.execute_script("arguments[0].scrollIntoView(true);", desc_element)
        desc_element.click()
        actions.send_keys(tab["Description"]).perform()

    # Availability
    print("Configurando la disponibilidad.")
    avail_elem = driver.find_element(
        By.XPATH, "//label[contains(@aria-label,'Disponibilidad')]"
    )
    driver.execute_script("arguments[0].scrollIntoView(true);", avail_elem)
    time.sleep(0.2)
    avail_elem.click()
    time.sleep(1)
    elements = driver.find_elements(By.XPATH, "//div[contains(@role,'option')]")
    for elem in elements:
        try:
            if "list as in stock" in str(elem.text.strip()).lower():
                elem.click()
                print("Disponibilidad configurada como 'in stock'")
                break
        except:
            pass

    # Location
    try:
        print("Buscando elemento de ubicación.")
        loc_element = driver.find_element(
            By.XPATH, "//label[contains(@aria-label,'Ubicación')]"
        )
    except:
        print("Elemento de ubicación no encontrado. Intentando usar 'more details'.")
        driver.execute_script(
            "arguments[0].scrollIntoView(true);", more_details_element
        )
        more_details_element.click()
        time.sleep(0.5)
        loc_element = driver.find_element(
            By.XPATH, "//label[contains(@aria-label,'Ubicación')]"
        )
    driver.execute_script("arguments[0].scrollIntoView(true);", loc_element)
    time.sleep(0.2)
    loc_element.click()
    actions.send_keys(tab["Location"]).perform()
    time.sleep(2)
    try:
        location_list_parent = driver.find_element(
            By.XPATH, "//ul[contains(@role,'listbox')]"
        )
        location_list = location_list_parent.find_elements(
            By.XPATH, "//li[contains(@role,'option')]"
        )
        location_list[0].click()
        print("Ubicación seleccionada.")
    except:
        print("Ubicación no seleccionada en el primer intento, intentando nuevamente.")
        time.sleep(3)
        location_list_parent = driver.find_element(
            By.XPATH, "//ul[contains(@role,'listbox')]"
        )
        location_list = location_list_parent.find_elements(
            By.XPATH, "//li[contains(@role,'option')]"
        )
        location_list[0].click()

    # Checkboxes
    try:
        print("Configurando las casillas de verificación.")
        cells = driver.find_elements(By.XPATH, "//div[contains(@role,'checkbox')]")
        for cell in cells:
            cell.click()
    except:
        print("No se encontraron casillas de verificación.")

    time.sleep(1)

    # Next
    print("Configurando el botón 'Siguiente'.")
    try:
        next_button = driver.find_element(By.XPATH, "//div[contains(@aria-label,'Siguiente')]")
        driver.execute_script("arguments[0].scrollIntoView(true);", next_button)
        next_button.click()
    except Exception as e:
        print(f"Error al hacer clic en el botón 'Siguiente': {e}")

    time.sleep(1)

    try:
        print("Buscando y haciendo clic en botones adicionales.")
        elements = driver.find_elements(By.XPATH, "//div[contains(@data-visualcompletion,'ignore-dynamic')]")
        
        total_buttons_clicked = 0
        max_buttons_to_click = 10  # Número máximo de botones a hacer clic
        
        for element in elements:
            if total_buttons_clicked >= max_buttons_to_click:
                break
            
            # Encuentra todos los div con role="button" dentro del elemento actual
            buttons = element.find_elements(By.XPATH, ".//div[@role='button']")
            
            for button in buttons:
                if total_buttons_clicked >= max_buttons_to_click:
                    break
                
                driver.execute_script("arguments[0].scrollIntoView(true);", button)
                button.click()
                total_buttons_clicked += 1
                print(f"Botón número {total_buttons_clicked} clickeado.")
                time.sleep(0.5)  # Ajusta el tiempo si es necesario entre clics

        if total_buttons_clicked >= max_buttons_to_click:
            # Solo haz clic en el botón de publicar si se ha hecho clic en la cantidad deseada de botones
            try:
                submit_button = driver.find_element(By.XPATH, "//div[contains(@aria-label,'Publicar')]")
                driver.execute_script("arguments[0].scrollIntoView(true);", submit_button)
                submit_button.click()
                print("Botón 'Publicar' clickeado.")
            except Exception as e:
                print(f"Error al hacer clic en el botón 'Publicar': {e}")

    except Exception as e:
        print(f"Error al encontrar o hacer clic en los botones: {e}")

    time.sleep(0.2)

def mark_as_sold(driver):
    try:
        print("Buscando y haciendo clic en 'Marcar como agotado' o 'Marcar como vendido'.")

        while True:
            # Encuentra todos los elementos con role="button" y aria-label que contengan "Marcar como agotado" o "Marcar como vendido"
            elements = driver.find_elements(By.XPATH, "//div[@role='button' and (contains(@aria-label,'Marcar como agotado') or contains(@aria-label,'Marcar como vendido'))]")
            
            # Si no hay más elementos para hacer clic, salir del bucle
            if not elements:
                print("No se encontraron más elementos.")
                break

            buttons_found = False  # Para verificar si se encontraron botones en la iteración actual
            
            for element in elements:
                aria_label = element.get_attribute("aria-label")
                driver.execute_script("arguments[0].scrollIntoView(true);", element)
                time.sleep(1.5)  # Espera 1 segundo antes de hacer clic
                
                try:
                    driver.execute_script("arguments[0].click();", element)
                    print(f"Botón '{aria_label}' clickeado.")
                    buttons_found = True
                    
                    if "Marcar como agotado" in aria_label:
                        # Espera y busca el botón con aria-label="Archivar"
                        time.sleep(2)
                        try:
                            archive_button = WebDriverWait(driver, 10).until(
                                EC.element_to_be_clickable((By.XPATH, "//div[@role='button' and @aria-label='Archivar']"))
                            )
                            driver.execute_script("arguments[0].scrollIntoView(true);", archive_button)
                            time.sleep(2)
                            archive_button.click()
                            print("Botón 'Archivar' clickeado.")
                        except Exception as archive_exception:
                            print(f"Error al hacer clic en el botón 'Archivar': {archive_exception}")

                    elif "Marcar como vendido" in aria_label:
                        # Espera y hace clic en el input con value="DECLINE"
                        time.sleep(2)
                        try:
                            decline_input = WebDriverWait(driver, 10).until(
                                EC.element_to_be_clickable((By.XPATH, "//input[@value='DECLINE']"))
                            )
                            decline_input.click()
                            print("Input 'DECLINE' clickeado.")
                            
                            # Espera y hace clic en el botón con aria-label="Siguiente"
                            next_button = WebDriverWait(driver, 10).until(
                                EC.element_to_be_clickable((By.XPATH, "//div[@role='button' and @aria-label='Siguiente']"))
                            )
                            driver.execute_script("arguments[0].scrollIntoView(true);", next_button)
                            time.sleep(2)
                            next_button.click()
                            print("Botón 'Siguiente' clickeado.")
                        except Exception as decline_exception:
                            print(f"Error al hacer clic en el input 'DECLINE' o el botón 'Siguiente': {decline_exception}")

                except Exception as click_exception:
                    print(f"Error al hacer clic en el botón: {click_exception}")

            if not buttons_found:
                print("No se encontraron más botones relevantes en la vista actual.")
                break
            
            # Realiza un scroll hacia abajo para cargar más elementos si es necesario
            driver.execute_script("window.scrollBy(0, window.innerHeight);")
            time.sleep(2)  # Espera a que la página cargue más elementos

        print("Proceso completado.")
    
    except Exception as e:
        print(f"Error general: {e}")

    time.sleep(0.2)



# Main Control Flow
if __name__ == "__main__":

        print("Cargando configuración.")
        config = read_settings()
        data = read_excel_tabs(config["tabs_file"])
        profiles = read_profiles(config["profiles_file"])
        if len(profiles) == 0:
            print("\n\nInsert Profiles in profiles.xlsx file to Run.\n")
            time.sleep(20)
            quit()
        print("\n")
        print("¿Que opción desea ejecutar?")
        print("1. Publicar Todas los articulos")
        print("2. Marcar como agotados/vendidos Todos los articulos")
        data = input("Seleccione un número: ")
        print("\n")

        if data == "1":
            if not images_exist(data):
                time.sleep(20)
                quit()
            drivers = []
            for profile in profiles:
                print(
                    "\nUsing Profile ", profile["Email"], " ->  Running for", len(data), "tabs."
                )
                driver = open_browser()
                actions = ActionChains(driver)
                driver.get("https://www.facebook.com/")
                time.sleep(5)
                login_facebook(driver, actions, profile)
                time.sleep(20)
                open_tabs(driver, len(data), 1)
                time.sleep(3)
                # Fill data
                for i in range(0, len(data)):
                    print("Filling Data for Tab#", str(i + 1))
                    driver.switch_to.window(driver.window_handles[i])
                    time.sleep(0.5)
                    fill_data(driver, actions, data[i])
                submit_quickly(driver)
                time.sleep(25)
                driver.quit()

        elif data == "2":
             for profile in profiles:
                print(
                    "\nUsing Profile ", profile["Email"], " ->  Running for", len(data), "tabs."
                )
                driver = open_browser()
                actions = ActionChains(driver)
                driver.get("https://www.facebook.com/")
                time.sleep(5)
                login_facebook(driver, actions, profile)
                time.sleep(20)
                open_tabs(driver, 1, 2)
                time.sleep(0.5)
                mark_as_sold(driver)
                time.sleep(25)
                driver.quit()
